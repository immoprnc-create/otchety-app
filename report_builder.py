"""
Report Builder — автосборка еженедельного отчёта из DataLens-выгрузок.
Зависимости: pip install customtkinter pandas openpyxl tkcalendar matplotlib
"""

import json
import os
import subprocess
import sys
import threading
from datetime import date, timedelta
from datetime import datetime as dt
from pathlib import Path

# ─── Проверка зависимостей ───────────────────────────────────────────────────

_missing = []
for _pkg in ("customtkinter", "pandas", "openpyxl", "tkcalendar", "matplotlib"):
    try:
        __import__(_pkg)
    except ImportError:
        _missing.append(_pkg)

if _missing:
    import tkinter as tk
    from tkinter import messagebox
    _r = tk.Tk(); _r.withdraw()
    messagebox.showerror(
        "Отсутствуют пакеты",
        f"Выполни в терминале:\n\npip install {' '.join(_missing)}"
    )
    sys.exit(1)

import customtkinter as ctk
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties, Outline as OutlineProperties
from tkcalendar import Calendar

import matplotlib
matplotlib.use("Agg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.ticker as ticker

# ─── Пути ────────────────────────────────────────────────────────────────────

if getattr(sys, "frozen", False):
    _APP_DIR = Path(sys.executable).parent
else:
    _APP_DIR = Path(__file__).parent

CONFIG_FILE       = _APP_DIR / "config.json"
HISTORY_FILE_NAME = "История.xlsx"

# ─── Цвета ───────────────────────────────────────────────────────────────────

_RED       = "#FC3F1D"
_RED_HOVER = "#E8381A"
_WHITE     = "#FFFFFF"
_DARK      = "#080808"
_BG        = "#111111"
_CARD      = "#1A1A1A"
_LABEL     = "#F0F0F0"
_SUBLABEL  = "#666666"
_BORDER    = "#272727"
_GHOST     = "#222222"
_GHOST_HOV = "#2E2E2E"
_FIELD     = "#141414"

# ─── Конфигурация показателей ────────────────────────────────────────────────

METRICS = [
    {
        "name": "Количество периодов",
        "prefix": "period_view_table",
        "employee_col": "Исполнитель",
        "value_col": None,
        "agg": "count",
        "pct": False,
        "csat_default": False,
    },
    {
        "name": "Количество уникальных",
        "prefix": "manager_actions_on_iterations",
        "employee_col": "Менеджер",
        "value_col": "Кол-во различных тикетов",
        "agg": "sum",
        "pct": False,
        "csat_default": False,
    },
    {
        "name": "Ёмкость",
        "prefix": "novoice_utilization",
        "employee_col": "ФИ",
        "value_col": "Capacity net на итерациях",
        "agg": "first",
        "pct": False,
        "csat_default": False,
    },
    {
        "name": "Утилизация %",
        "prefix": "novoice_utilization",
        "employee_col": "ФИ",
        "value_col": "Utilization",
        "agg": "first",
        "pct": True,
        "csat_default": False,
    },
    {
        "name": "Окупанси %",
        "prefix": "novoice_utilization",
        "employee_col": "ФИ",
        "value_col": "Occupancy",
        "agg": "first",
        "pct": True,
        "csat_default": False,
    },
    {
        "name": "Время обработки",
        "prefix": "queue_period_view_table",
        "employee_col": "Менеджер",
        "value_col": "Ср. время обработки оператора (мин)",
        "agg": "mean",
        "pct": False,
        "csat_default": False,
    },
    {
        "name": "Время реакции",
        "prefix": "queue_period_view_table",
        "employee_col": "Менеджер",
        "value_col": "Ср. время реакции менеджера (мин)",
        "agg": "mean",
        "pct": False,
        "csat_default": False,
    },
    {
        "name": "Время решения",
        "prefix": "queue_period_view_table",
        "employee_col": "Менеджер",
        "value_col": "Ср. время решения оператора (мин)",
        "agg": "mean",
        "pct": False,
        "csat_default": False,
    },
    {
        "name": "CSAT финальный",
        "prefix": "orate_votes_view_table",
        "csat_type": "Финальная оценка",
        "employee_col": "Сотрудник",
        "value_col": "Оценка, балл",
        "agg": "mean",
        "pct": False,
        "csat_default": True,
    },
    {
        "name": "CSAT промеж. (ср. балл)",
        "prefix": "orate_votes_view_table",
        "csat_type": "Промежуточная оценка",
        "employee_col": "Сотрудник",
        "value_col": "Оценка, балл",
        "agg": "mean",
        "pct": False,
        "csat_default": True,
    },
    {
        "name": "Кол-во тикетов на проверку",
        "prefix": "основная информация качество",
        "employee_col": "Оператор",
        "value_col": None,
        "agg": "count",
        "pct": False,
        "csat_default": False,
        "zero_default": True,
    },
    {
        "name": "Критичные ошибки",
        "prefix": "основная информация качество",
        "employee_col": "Оператор",
        "value_col": "Крит ошибка КПИ",
        "agg": "sum",
        "pct": False,
        "csat_default": False,
        "zero_default": True,
    },
    {
        "name": "Некритичные ошибки",
        "prefix": "основная информация качество",
        "employee_col": "Оператор",
        "value_col": "Не крит ошибка КПИ",
        "agg": "sum",
        "pct": False,
        "csat_default": False,
        "zero_default": True,
    },
]

CSAT_REASON_COL = "Причина"
CSAT_TICKET_COL = "Ссылка на тикет"

_RANKING_METRICS = ["Ёмкость", "CSAT финальный", "Критичные ошибки"]
_TOP_CRITERIA = {
    "Ёмкость":          lambda v: v >= 3.0,
    "CSAT финальный":   lambda v: v >= 4.4,
    "Критичные ошибки": lambda v: v == 0,
}
_ANTI_TOP_CRITERIA = {
    "Ёмкость":          lambda v: v <= 2.99,
    "CSAT финальный":   lambda v: v <= 4.39,
    "Критичные ошибки": lambda v: v >= 2,
}
_TOP_CRITERIA_TEXT = {
    "Ёмкость": "≥ 3.0", "CSAT финальный": "≥ 4.4", "Критичные ошибки": "= 0",
}
_ANTI_TOP_CRITERIA_TEXT = {
    "Ёмкость": "≤ 2.99", "CSAT финальный": "≤ 4.39", "Критичные ошибки": "≥ 2",
}
_TOP_SORT_REVERSE = {
    "Ёмкость": True, "CSAT финальный": True, "Критичные ошибки": False,
}
_ANTI_TOP_SORT_REVERSE = {
    "Ёмкость": False, "CSAT финальный": False, "Критичные ошибки": True,
}

# ─── Конфиг ──────────────────────────────────────────────────────────────────

def _load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def _save_config(data: dict):
    try:
        CONFIG_FILE.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception:
        pass

# ─── Вспомогательные ─────────────────────────────────────────────────────────

def _auto_period() -> str:
    today = date.today()
    days_since_mon = today.weekday()
    last_mon = today - timedelta(days=days_since_mon + 7)
    last_sun = last_mon + timedelta(days=6)
    return f"{last_mon.day:02d}.{last_mon.month:02d}-{last_sun.day:02d}.{last_sun.month:02d}"

def _open_path(path):
    if sys.platform == "win32":
        os.startfile(str(path))
    else:
        subprocess.run(["open", str(path)])

def _notify(title: str, body: str):
    if sys.platform == "win32":
        try:
            script = (
                "Add-Type -AssemblyName System.Windows.Forms;"
                "$n=New-Object System.Windows.Forms.NotifyIcon;"
                "$n.Icon=[System.Drawing.SystemIcons]::Application;"
                "$n.Visible=$true;"
                f'$n.ShowBalloonTip(5000,"{title}","{body}",'
                "[System.Windows.Forms.ToolTipIcon]::None);"
                "Start-Sleep 6;$n.Dispose()"
            )
            subprocess.Popen(
                ["powershell", "-WindowStyle", "Hidden", "-NoProfile", "-Command", script],
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
        except Exception:
            pass
    elif sys.platform == "darwin":
        try:
            subprocess.Popen(
                ["osascript", "-e", f'display notification "{body}" with title "{title}"']
            )
        except Exception:
            pass

# ─── Поиск файлов ────────────────────────────────────────────────────────────

def _find_xlsx_by_prefix(folder: Path, prefix: str) -> Path:
    prefix_low = prefix.lower()
    matches = [f for f in folder.rglob("*.xlsx") if f.name.lower().startswith(prefix_low)]
    if not matches:
        raise FileNotFoundError(f"Файл с началом «{prefix}» не найден")
    return matches[0]

def _find_csat_xlsx(folder: Path, csat_type: str) -> Path:
    matches = [f for f in folder.rglob("*.xlsx")
               if f.name.lower().startswith("orate_votes_view_table")]
    if not matches:
        raise FileNotFoundError("CSAT файл не найден")
    for path in matches:
        try:
            df = pd.read_excel(path, engine="openpyxl", nrows=5)
            if "Тип оценки" in df.columns:
                vals = df["Тип оценки"].dropna()
                if not vals.empty and csat_type.lower() in str(vals.iloc[0]).lower():
                    return path
        except Exception:
            pass
    raise FileNotFoundError(f"CSAT файл с типом «{csat_type}» не найден")

# ─── Обработка данных ────────────────────────────────────────────────────────

def _to_numeric_binary(series: pd.Series) -> pd.Series:
    mapped = series.map({"Да": 1, "Нет": 0, "да": 1, "нет": 0, "ДА": 1, "НЕТ": 0})
    result = mapped.where(mapped.notna(), series)
    return pd.to_numeric(result, errors="coerce").fillna(0)

def _read_metric(folder: Path, cfg: dict) -> pd.Series:
    if "csat_type" in cfg:
        xlsx = _find_csat_xlsx(folder, cfg["csat_type"])
    else:
        xlsx = _find_xlsx_by_prefix(folder, cfg["prefix"])

    df = pd.read_excel(xlsx, engine="openpyxl")

    if "csat_type" in cfg and "Тип оценки" in df.columns:
        df = df[df["Тип оценки"].astype(str).str.strip() == cfg["csat_type"]]

    emp = cfg["employee_col"]
    val = cfg["value_col"]
    agg = cfg["agg"]

    if emp not in df.columns:
        raise KeyError(f"Столбец «{emp}» не найден в «{xlsx.name}»")

    df = df[df[emp].notna() & (df[emp].astype(str).str.strip() != "")]

    if val is None:
        return df.groupby(emp, sort=True).size().rename(cfg["name"])

    if val not in df.columns:
        raise KeyError(f"Столбец «{val}» не найден в «{xlsx.name}»")

    if agg == "sum":
        df = df.copy()
        df[val] = _to_numeric_binary(df[val])
        return df.groupby(emp, sort=True)[val].sum().rename(cfg["name"])
    elif agg == "mean":
        return df.groupby(emp, sort=True)[val].mean().rename(cfg["name"])
    else:
        return df.groupby(emp, sort=True)[val].first().rename(cfg["name"])

def _read_csat_types(folder: Path) -> pd.DataFrame | None:
    try:
        xlsx = _find_csat_xlsx(folder, "Промежуточная оценка")
    except FileNotFoundError:
        return None
    df = pd.read_excel(xlsx, engine="openpyxl")
    if "Тип оценки" in df.columns:
        df = df[df["Тип оценки"].astype(str).str.strip() == "Промежуточная оценка"]
    if CSAT_REASON_COL not in df.columns or CSAT_TICKET_COL not in df.columns:
        return None
    counts = (
        df.groupby(CSAT_REASON_COL)[CSAT_TICKET_COL]
        .count().reset_index()
    )
    counts.columns = ["Причина", "Количество"]
    counts = (
        counts[counts["Количество"] > 0]
        .sort_values("Количество", ascending=False)
        .reset_index(drop=True)
    )
    return counts if not counts.empty else None

def scan_folder(cifry_folder: str) -> dict:
    base = Path(cifry_folder)
    names: set[str] = set()
    found_keys: set[str] = set()
    found = 0
    missing = []

    for cfg in METRICS:
        key = cfg.get("csat_type") or cfg["prefix"]
        file_found = False

        if key not in found_keys:
            try:
                if "csat_type" in cfg:
                    _find_csat_xlsx(base, cfg["csat_type"])
                else:
                    _find_xlsx_by_prefix(base, cfg["prefix"])
                found += 1
                found_keys.add(key)
                file_found = True
            except FileNotFoundError:
                missing.append(cfg["name"])
        else:
            file_found = True

        if not file_found:
            continue
        try:
            if "csat_type" in cfg:
                xlsx = _find_csat_xlsx(base, cfg["csat_type"])
                df = pd.read_excel(xlsx, engine="openpyxl")
                if "Тип оценки" in df.columns:
                    df = df[df["Тип оценки"].astype(str).str.strip() == cfg["csat_type"]]
            else:
                xlsx = _find_xlsx_by_prefix(base, cfg["prefix"])
                df = pd.read_excel(xlsx, engine="openpyxl", usecols=[cfg["employee_col"]])
            vals = df[cfg["employee_col"]].dropna().astype(str).str.strip()
            names.update(v for v in vals if v)
        except Exception:
            pass

    return {"employees": sorted(names), "found": found, "missing": missing}

def collect_employees(cifry_folder: str) -> list[str]:
    return scan_folder(cifry_folder)["employees"]

# ─── Сборка отчёта ───────────────────────────────────────────────────────────

def build_report(
    cifry_folder: str,
    period: str,
    employee_filter: set[str] | None,
    fired_employees: set[str],
    progress_cb,
    status_cb,
) -> str:
    base = Path(cifry_folder)
    series_list = []
    errors = []
    total = len(METRICS)

    for i, cfg in enumerate(METRICS):
        status_cb(f"Обрабатывается: {cfg['name']}…")
        try:
            s = _read_metric(base, cfg)
            if cfg["pct"]:
                s = (s * 100).round(2)
            series_list.append(s)
        except Exception as e:
            errors.append(f"{cfg['name']}: {e}")
        progress_cb((i + 1) / total)

    if not series_list:
        raise RuntimeError("Не удалось обработать ни один показатель.\n\n" + "\n".join(errors))

    status_cb("Формирование сводной таблицы…")
    summary = pd.concat(series_list, axis=1)
    summary.index.name = "Сотрудник"
    summary = summary.sort_index()

    for cfg in METRICS:
        if cfg["csat_default"] and cfg["name"] in summary.columns:
            summary[cfg["name"]] = summary[cfg["name"]].fillna(5)
        if cfg.get("zero_default") and cfg["name"] in summary.columns:
            summary[cfg["name"]] = summary[cfg["name"]].fillna(0)

    if employee_filter is not None:
        summary = summary[summary.index.isin(employee_filter)]

    status_cb("Обновление истории…")
    try:
        _append_to_history(summary, period, base.parent, fired_employees)
    except Exception as e:
        errors.append(f"История: {e}")

    return str(base.parent / HISTORY_FILE_NAME)

# ─── Накопительный файл История.xlsx ─────────────────────────────────────────
#
# Формат:
#   Строка 1:  "Показатели" | Период1 | Период2 | ...   (заголовки столбцов)
#   Далее блоки на каждую метрику:
#     Строка-заголовок метрики: название | итог_п1 | итог_п2 | ...
#     Строки сотрудников:       имя      | знач_п1 | знач_п2 | ...  (outline_level=1)
#   Столбцы периодов — outline_level=1 (сворачиваются через плюсик сверху).
#   Строки сотрудников — outline_level=1 (сворачиваются через плюсик слева).

def _read_history_data(
    hist_path: Path,
) -> tuple[list[str], dict[str, dict[str, dict[str, object]]]]:
    metric_names = [cfg["name"] for cfg in METRICS]
    data: dict[str, dict[str, dict[str, object]]] = {m: {} for m in metric_names}
    periods: list[str] = []

    if not hist_path.exists():
        return periods, data

    wb = openpyxl.load_workbook(hist_path, read_only=True, data_only=True)
    ws = wb.active
    max_col = ws.max_column or 1
    max_row = ws.max_row or 0

    for col in range(2, max_col + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            periods.append(str(v).strip())

    current_metric = None
    for row in range(2, max_row + 1):
        a_val = ws.cell(row=row, column=1).value
        if not a_val:
            continue
        a_str = str(a_val).strip()
        if a_str in metric_names:
            current_metric = a_str
        elif current_metric:
            emp = a_str
            if emp not in data[current_metric]:
                data[current_metric][emp] = {}
            for pi, p in enumerate(periods):
                val = ws.cell(row=row, column=2 + pi).value
                if val not in (None, "—"):
                    data[current_metric][emp][p] = val

    wb.close()
    return periods, data


def _write_history_full(
    hist_path: Path,
    periods: list[str],
    data: dict[str, dict[str, dict[str, object]]],
):
    metric_names = [cfg["name"] for cfg in METRICS]
    metric_agg   = {cfg["name"]: cfg["agg"] for cfg in METRICS}

    DARK_HDR  = PatternFill("solid", fgColor="111827")
    PER_FILL  = PatternFill("solid", fgColor="FC3F1D")
    MET_FILL  = PatternFill("solid", fgColor="1F2937")
    ALT1_FILL = PatternFill("solid", fgColor="F5F5F5")
    ALT2_FILL = PatternFill("solid", fgColor="FFFFFF")

    HDR_F  = Font(bold=True, color="FFFFFF", size=11)
    MET_F  = Font(bold=True, color="FFFFFF", size=10)
    NAME_F = Font(bold=True, color="111827", size=10)
    VAL_F  = Font(color="111827", size=10)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История"
    ws.sheet_properties = WorksheetProperties(
        outlinePr=OutlineProperties(summaryBelow=False, summaryRight=False)
    )

    # Строка 1: заголовки
    ws.column_dimensions["A"].width = 28
    ws.row_dimensions[1].height = 26
    c = ws.cell(row=1, column=1, value="Показатели")
    c.fill = DARK_HDR; c.font = HDR_F; c.alignment = LEFT

    for pi, p in enumerate(periods):
        col = 2 + pi
        c = ws.cell(row=1, column=col, value=p)
        c.fill = PER_FILL; c.font = HDR_F; c.alignment = CENTER
        ltr = get_column_letter(col)
        ws.column_dimensions[ltr].width = 14
        ws.column_dimensions[ltr].outline_level = 1

    # Блоки метрик
    current_row = 2
    for metric in metric_names:
        emp_dict = data.get(metric, {})
        all_employees = sorted(emp_dict.keys())
        agg = metric_agg.get(metric, "mean")

        # Строка-заголовок метрики с итогами
        ws.row_dimensions[current_row].height = 24
        c = ws.cell(row=current_row, column=1, value=metric)
        c.fill = MET_FILL; c.font = MET_F; c.alignment = LEFT

        for pi, p in enumerate(periods):
            col = 2 + pi
            vals = []
            for e in all_employees:
                v = emp_dict[e].get(p)
                if v is not None and v != "—":
                    try:
                        vals.append(float(v))
                    except (TypeError, ValueError):
                        pass
            if vals:
                raw = sum(vals) if agg in ("sum", "count") else sum(vals) / len(vals)
                total = int(raw) if raw == int(raw) else round(raw, 2)
            else:
                total = "—"
            c = ws.cell(row=current_row, column=col, value=total)
            c.fill = MET_FILL; c.font = MET_F; c.alignment = CENTER

        current_row += 1

        # Строки сотрудников (сворачиваемые)
        for ri, emp in enumerate(all_employees):
            fill = ALT1_FILL if ri % 2 == 0 else ALT2_FILL
            c = ws.cell(row=current_row, column=1, value=emp)
            c.fill = fill; c.font = NAME_F; c.alignment = LEFT

            for pi, p in enumerate(periods):
                col = 2 + pi
                val = emp_dict[emp].get(p)
                if val is None:
                    display = "—"
                elif isinstance(val, float):
                    display = round(val, 2)
                else:
                    display = val
                c2 = ws.cell(row=current_row, column=col, value=display)
                c2.fill = fill; c2.font = VAL_F; c2.alignment = CENTER

            ws.row_dimensions[current_row].outline_level = 1
            current_row += 1

    ws.freeze_panes = "B2"
    wb.save(hist_path)


def _append_to_history(summary: pd.DataFrame, period: str, out_dir: Path, fired: set[str]):
    hist_path = out_dir / HISTORY_FILE_NAME

    periods, data = _read_history_data(hist_path)

    # Удаляем уволенных
    for metric in data:
        for emp in list(fired):
            data[metric].pop(emp, None)

    # Добавляем данные нового периода
    if period not in periods:
        periods.append(period)

    for cfg in METRICS:
        metric = cfg["name"]
        if metric not in summary.columns:
            continue
        if metric not in data:
            data[metric] = {}
        for emp in summary.index:
            val = summary.loc[emp, metric]
            if emp not in data[metric]:
                data[metric][emp] = {}
            data[metric][emp][period] = None if pd.isna(val) else val

    _write_history_full(hist_path, periods, data)

# ─── Чтение истории (для Динамики и Топ/Анти Топ) ────────────────────────────

def get_history_periods(hist_path: Path) -> list[str]:
    periods, _ = _read_history_data(hist_path)
    return periods


def load_history_for_chart(hist_path: Path, employee: str, metric: str) -> pd.Series:
    if not hist_path.exists():
        return pd.Series(dtype=float)
    periods, data = _read_history_data(hist_path)
    emp_data = data.get(metric, {}).get(employee, {})
    result = {}
    for p in periods:
        val = emp_data.get(p)
        if val is not None and val != "—":
            try:
                result[p] = float(val)
            except (TypeError, ValueError):
                pass
    return pd.Series(result)


def load_last_period_summary(hist_path: Path) -> pd.DataFrame | None:
    if not hist_path.exists():
        return None
    periods, data = _read_history_data(hist_path)
    if not periods:
        return None
    last = periods[-1]
    metric_names = [cfg["name"] for cfg in METRICS]

    all_employees: set[str] = set()
    for m in metric_names:
        all_employees.update(data.get(m, {}).keys())
    if not all_employees:
        return None

    rows = {}
    for emp in all_employees:
        row = {}
        for m in metric_names:
            val = data.get(m, {}).get(emp, {}).get(last)
            if val is not None and val != "—":
                try:
                    row[m] = float(val)
                except (TypeError, ValueError):
                    row[m] = None
            else:
                row[m] = None
        rows[emp] = row

    df = pd.DataFrame(rows).T
    df.index.name = "Сотрудник"
    return df


def delete_period_from_history(hist_path: Path, period: str):
    if not hist_path.exists():
        return
    periods, data = _read_history_data(hist_path)
    if period not in periods:
        return
    periods.remove(period)
    for metric in data:
        for emp in data[metric]:
            data[metric][emp].pop(period, None)
    _write_history_full(hist_path, periods, data)

# ─── Запись Excel (вертикальный формат, без ТОП/Антитоп) ─────────────────────

def _write_excel(
    path: Path,
    summary: pd.DataFrame,
    csat_types,
    period: str,
    errors: list,
    totals: dict,
):
    HDR_FILL  = PatternFill("solid", fgColor="FC3F1D")
    COL_FILL  = PatternFill("solid", fgColor="1F3864")
    ALT_FILL  = PatternFill("solid", fgColor="F5F5F5")
    WHITE     = PatternFill("solid", fgColor="FFFFFF")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=12)
    COL_FONT  = Font(bold=True, color="FFFFFF", size=10)
    EMP_FONT  = Font(bold=True, color="000000", size=10)
    CELL_FONT = Font(color="000000", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сводная"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    current_row = 1

    for col_name in summary.columns:
        ws.merge_cells(f"A{current_row}:B{current_row}")
        c = ws.cell(row=current_row, column=1, value=col_name)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = LEFT
        ws.cell(row=current_row, column=2).fill = HDR_FILL
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        total_raw = totals.get(col_name)
        total_label = str(total_raw) if total_raw is not None else "—"
        for ci, label in enumerate(["Сотрудник", total_label], 1):
            c = ws.cell(row=current_row, column=ci, value=label)
            c.fill = COL_FILL; c.font = COL_FONT; c.alignment = CENTER
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for ri, (emp, val) in enumerate(summary[col_name].items()):
            fill = ALT_FILL if ri % 2 == 0 else WHITE
            ec = ws.cell(row=current_row, column=1, value=emp)
            ec.fill = fill; ec.font = EMP_FONT; ec.alignment = LEFT
            if pd.isna(val):
                display = "—"
            elif isinstance(val, float):
                display = round(val, 2)
            else:
                display = val
            vc = ws.cell(row=current_row, column=2, value=display)
            vc.fill = fill; vc.font = CELL_FONT; vc.alignment = CENTER
            current_row += 1

        for ci in (1, 2):
            ws.cell(row=current_row, column=ci).fill = WHITE
        current_row += 1

        if col_name == "CSAT промеж. (ср. балл)" and csat_types is not None:
            total_ct = int(csat_types["Количество"].sum())
            ws.merge_cells(f"A{current_row}:B{current_row}")
            c = ws.cell(row=current_row, column=1, value="CSAT промеж. — Причины")
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = LEFT
            ws.cell(row=current_row, column=2).fill = HDR_FILL
            ws.row_dimensions[current_row].height = 28
            current_row += 1
            for ci, lbl in enumerate(["Причина", str(total_ct)], 1):
                c = ws.cell(row=current_row, column=ci, value=lbl)
                c.fill = COL_FILL; c.font = COL_FONT; c.alignment = CENTER
            ws.row_dimensions[current_row].height = 22
            current_row += 1
            for ri, row in enumerate(csat_types.itertuples(index=False)):
                fill = ALT_FILL if ri % 2 == 0 else WHITE
                c1 = ws.cell(row=current_row, column=1, value=row[0])
                c1.fill = fill; c1.font = EMP_FONT; c1.alignment = LEFT
                c2 = ws.cell(row=current_row, column=2, value=int(row[1]))
                c2.fill = fill; c2.font = CELL_FONT; c2.alignment = CENTER
                current_row += 1
            for ci in (1, 2):
                ws.cell(row=current_row, column=ci).fill = WHITE
            current_row += 1

    for ri in range(current_row, current_row + 60):
        for ci in (1, 2):
            ws.cell(row=ri, column=ci).fill = WHITE

    if errors:
        ws3 = wb.create_sheet("⚠ Ошибки")
        ws3.cell(row=1, column=1, value="Ошибки при обработке").font = Font(
            bold=True, color="C00000", size=12
        )
        for ri, err in enumerate(errors, 2):
            ws3.cell(row=ri, column=1, value=err).font = Font(size=10, color="000000")
        ws3.column_dimensions["A"].width = 90

    wb.save(path)

# ─── Окно фильтра сотрудников ────────────────────────────────────────────────

class FilterWindow(ctk.CTkToplevel):
    _REASONS = ["Отпуск", "Больничный", "Уволился", "РГ"]

    def __init__(self, parent, employees: list[str], scan_result: dict,
                 excluded: set[str], on_confirm):
        super().__init__(parent)
        self.title("Сотрудники")
        self.geometry("440x580")
        self.resizable(False, False)
        self.configure(fg_color=_BG)
        self.grab_set()
        self._on_confirm = on_confirm
        self._rows: dict[str, dict] = {}
        self._build_ui(employees, scan_result, excluded)

    def _build_ui(self, employees, scan_result, excluded):
        header = ctk.CTkFrame(self, fg_color=_DARK, corner_radius=0, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(
            header, text="Выбор сотрудников",
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            text_color=_WHITE,
        ).pack(side="left", padx=20)

        n_found = scan_result.get("found", len(METRICS))
        missing = scan_result.get("missing", [])
        n_emp   = len(employees)
        total_files = len({cfg.get("csat_type") or cfg["prefix"] for cfg in METRICS})
        if missing:
            info_color = "#D97706"
            info_text  = f"✓  {n_emp} сотрудников   ·   {n_found}/{total_files} файлов   ⚠  нет: {', '.join(missing[:2])}"
            if len(missing) > 2:
                info_text += f" и ещё {len(missing)-2}"
        else:
            info_color = "#059669"
            info_text  = f"✓  {n_emp} сотрудников   ·   все {total_files} файла найдены"

        ctk.CTkLabel(self, text=info_text,
                     font=ctk.CTkFont(family="Segoe UI", size=11),
                     text_color=info_color, anchor="w").pack(fill="x", padx=20, pady=(10, 2))
        ctk.CTkLabel(self,
                     text="Снимите галочку и выберите причину отсутствия",
                     font=ctk.CTkFont(family="Segoe UI", size=12),
                     text_color=_SUBLABEL, anchor="w").pack(fill="x", padx=20, pady=(0, 6))

        self._search_var = ctk.StringVar()
        ctk.CTkEntry(
            self, textvariable=self._search_var,
            placeholder_text="Поиск сотрудника…",
            height=34, border_color=_BORDER, border_width=1,
            fg_color=_FIELD, text_color=_LABEL, corner_radius=8,
            font=ctk.CTkFont(family="Segoe UI", size=12),
        ).pack(fill="x", padx=20, pady=(0, 6))
        self._search_var.trace_add("write", self._on_search)

        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(0, 8))
        for txt, cmd in [("Выбрать всех", self._select_all), ("Снять всех", self._deselect_all)]:
            ctk.CTkButton(btn_row, text=txt, height=30, width=110,
                          fg_color=_GHOST, hover_color=_GHOST_HOV,
                          text_color=_LABEL, corner_radius=8,
                          border_width=1, border_color=_BORDER,
                          font=ctk.CTkFont(family="Segoe UI", size=12),
                          command=cmd).pack(side="left", padx=(0, 8))

        scroll = ctk.CTkScrollableFrame(self, fg_color=_CARD,
                                        border_color=_BORDER, border_width=1,
                                        corner_radius=10)
        scroll.pack(fill="both", expand=True, padx=20, pady=(0, 12))

        for emp in employees:
            active_var = ctk.BooleanVar(value=(emp not in excluded))
            reason_var = ctk.StringVar(value="Отпуск")
            row_frame  = ctk.CTkFrame(scroll, fg_color="transparent")
            row_frame.pack(fill="x", pady=2, padx=4)

            ctk.CTkCheckBox(
                row_frame, text=emp, variable=active_var,
                font=ctk.CTkFont(family="Segoe UI", size=12),
                text_color=_LABEL, fg_color=_RED, hover_color=_RED_HOVER,
                checkmark_color=_WHITE, border_color=_BORDER,
            ).pack(side="left")

            menu = ctk.CTkOptionMenu(
                row_frame, values=self._REASONS, variable=reason_var,
                width=118, height=28,
                fg_color=_GHOST, button_color=_GHOST, button_hover_color=_GHOST_HOV,
                text_color=_LABEL, font=ctk.CTkFont(family="Segoe UI", size=11),
                command=lambda val, e=emp: self._on_reason_change(e, val),
            )
            if not active_var.get():
                menu.pack(side="right")

            self._rows[emp] = {"active": active_var, "reason": reason_var, "menu": menu, "frame": row_frame}
            active_var.trace_add("write", lambda *_, e=emp: self._on_active_change(e))

        ctk.CTkButton(self, text="Сформировать отчёт", height=46,
                      font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
                      fg_color=_RED, hover_color=_RED_HOVER,
                      text_color=_WHITE, corner_radius=12,
                      command=self._confirm).pack(fill="x", padx=20, pady=(0, 16))

    def _on_active_change(self, emp):
        row = self._rows[emp]
        if row["active"].get():
            row["menu"].pack_forget(); row["reason"].set("Отпуск")
        else:
            row["menu"].pack(side="right")

    def _on_reason_change(self, emp, val):
        if val != "Уволился":
            return
        dlg = ctk.CTkToplevel(self)
        dlg.title(""); dlg.geometry("340x170")
        dlg.resizable(False, False)
        dlg.configure(fg_color=_BG)
        dlg.grab_set(); dlg.after(50, dlg.lift)
        ctk.CTkLabel(dlg, text=f"Отметить «{emp}» как уволенного?",
                     font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                     text_color=_LABEL, wraplength=290).pack(pady=(22, 4), padx=24)
        ctk.CTkLabel(dlg, text="Его строка будет удалена из файла «История.xlsx»",
                     font=ctk.CTkFont(family="Segoe UI", size=11),
                     text_color=_SUBLABEL, wraplength=290).pack(pady=(0, 18), padx=24)
        br = ctk.CTkFrame(dlg, fg_color="transparent")
        br.pack(fill="x", padx=24)
        ctk.CTkButton(br, text="Да, уволился",
                      fg_color="#DC2626", hover_color="#B91C1C",
                      text_color=_WHITE, height=38, corner_radius=8,
                      font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                      command=dlg.destroy).pack(side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkButton(br, text="Отмена",
                      fg_color=_GHOST, hover_color=_GHOST_HOV,
                      text_color=_LABEL, height=38, corner_radius=8,
                      border_width=1, border_color=_BORDER,
                      font=ctk.CTkFont(family="Segoe UI", size=13),
                      command=lambda: (self._rows[emp]["reason"].set("Отпуск"), dlg.destroy()),
                      ).pack(side="left", fill="x", expand=True)

    def _on_search(self, *_):
        query = self._search_var.get().strip().lower()
        for row in self._rows.values():
            row["frame"].pack_forget()
        for emp, row in self._rows.items():
            if not query or query in emp.lower():
                row["frame"].pack(fill="x", pady=2, padx=4)

    def _select_all(self):
        for row in self._rows.values(): row["active"].set(True)

    def _deselect_all(self):
        for row in self._rows.values(): row["active"].set(False)

    def _confirm(self):
        selected, fired, excluded = set(), set(), []
        for name, row in self._rows.items():
            if row["active"].get():
                selected.add(name)
            elif row["reason"].get() == "Уволился":
                fired.add(name)
            else:
                excluded.append(name)
        cfg = _load_config()
        cfg["excluded_employees"] = excluded
        _save_config(cfg)
        self.destroy()
        self._on_confirm(selected, fired)

# ─── Окно управления историей ────────────────────────────────────────────────

class HistoryManagerWindow(ctk.CTkToplevel):
    def __init__(self, parent, hist_path: Path):
        super().__init__(parent)
        self.title("Управление историей")
        self.geometry("420x480")
        self.resizable(False, False)
        self.configure(fg_color=_BG)
        self.grab_set()
        self._hist_path = hist_path
        self._build_ui()

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=_DARK, corner_radius=0, height=60)
        header.pack(fill="x"); header.pack_propagate(False)
        ctk.CTkLabel(header, text="Управление историей",
                     font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
                     text_color=_WHITE).pack(side="left", padx=20)

        ctk.CTkLabel(self,
                     text="Нажмите корзину рядом с периодом, чтобы удалить его из Истории",
                     font=ctk.CTkFont(family="Segoe UI", size=11),
                     text_color=_SUBLABEL, anchor="w", wraplength=380,
                     ).pack(fill="x", padx=20, pady=(12, 8))

        self._list_frame = ctk.CTkScrollableFrame(
            self, fg_color=_CARD, border_color=_BORDER,
            border_width=1, corner_radius=10,
        )
        self._list_frame.pack(fill="both", expand=True, padx=20, pady=(0, 12))
        self._refresh_list()

        ctk.CTkButton(self, text="Открыть файл История.xlsx",
                      height=42, corner_radius=10,
                      fg_color=_GHOST, hover_color=_GHOST_HOV,
                      text_color=_LABEL, border_width=1, border_color=_BORDER,
                      font=ctk.CTkFont(family="Segoe UI", size=13),
                      command=lambda: _open_path(self._hist_path),
                      ).pack(fill="x", padx=20, pady=(0, 16))

    def _refresh_list(self):
        for w in self._list_frame.winfo_children():
            w.destroy()

        periods = get_history_periods(self._hist_path)
        if not periods:
            ctk.CTkLabel(self._list_frame, text="История пуста",
                         font=ctk.CTkFont(family="Segoe UI", size=12),
                         text_color=_SUBLABEL).pack(pady=20)
            return

        for i, period in enumerate(periods):
            row = ctk.CTkFrame(self._list_frame,
                               fg_color="#141414" if i % 2 == 0 else _CARD,
                               corner_radius=6, height=38)
            row.pack(fill="x", pady=2, padx=4)
            row.pack_propagate(False)
            ctk.CTkLabel(row, text=period,
                         font=ctk.CTkFont(family="Segoe UI", size=13),
                         text_color=_LABEL).pack(side="left", padx=12)
            ctk.CTkButton(row, text="🗑", width=32, height=28,
                          fg_color="transparent", hover_color="#FEE2E2",
                          text_color="#DC2626", corner_radius=6,
                          font=ctk.CTkFont(family="Segoe UI", size=14),
                          command=lambda p=period: self._ask_delete(p),
                          ).pack(side="right", padx=6)

    def _ask_delete(self, period: str):
        dlg = ctk.CTkToplevel(self)
        dlg.title(""); dlg.geometry("320x155")
        dlg.resizable(False, False)
        dlg.configure(fg_color=_BG)
        dlg.grab_set(); dlg.after(50, dlg.lift)

        ctk.CTkLabel(dlg, text=f"Удалить период «{period}»?",
                     font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                     text_color=_LABEL).pack(pady=(22, 4), padx=24)
        ctk.CTkLabel(dlg, text="Данные за этот период будут удалены из Истории",
                     font=ctk.CTkFont(family="Segoe UI", size=11),
                     text_color=_SUBLABEL, wraplength=270).pack(pady=(0, 16), padx=24)

        br = ctk.CTkFrame(dlg, fg_color="transparent")
        br.pack(fill="x", padx=24)

        def confirm():
            dlg.destroy()
            delete_period_from_history(self._hist_path, period)
            self._refresh_list()

        ctk.CTkButton(br, text="Удалить",
                      fg_color="#DC2626", hover_color="#B91C1C",
                      text_color=_WHITE, height=36, corner_radius=8,
                      font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                      command=confirm).pack(side="left", fill="x", expand=True, padx=(0, 6))
        ctk.CTkButton(br, text="Отмена",
                      fg_color=_GHOST, hover_color=_GHOST_HOV,
                      text_color=_LABEL, height=36, corner_radius=8,
                      border_width=1, border_color=_BORDER,
                      font=ctk.CTkFont(family="Segoe UI", size=13),
                      command=dlg.destroy).pack(side="left", fill="x", expand=True)


# ─── Окно выбора периода (календарь) ─────────────────────────────────────────

class DatePickerWindow(ctk.CTkToplevel):
    def __init__(self, parent, on_confirm):
        super().__init__(parent)
        self.title("Выбор периода")
        self.resizable(False, False)
        self.configure(fg_color=_BG)
        self.grab_set()
        self._on_confirm = on_confirm
        self._build_ui()
        self.update_idletasks()
        self.geometry(f"{self.winfo_reqwidth()}x{self.winfo_reqheight()}")

    def _build_ui(self):
        header = ctk.CTkFrame(self, fg_color=_DARK, corner_radius=0, height=60)
        header.pack(fill="x"); header.pack_propagate(False)
        ctk.CTkLabel(header, text="Выбор периода",
                     font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
                     text_color=_WHITE).pack(side="left", padx=20)

        cal_kwargs = dict(
            selectmode="day", foreground=_LABEL, background=_CARD,
            disabledbackground=_CARD, bordercolor=_BORDER,
            headersbackground=_RED, headersforeground=_WHITE,
            selectbackground=_RED, selectforeground=_WHITE,
            normalbackground=_CARD, normalforeground=_LABEL,
            weekendbackground=_CARD, weekendforeground=_RED,
            othermonthbackground=_BG, othermonthforeground="#4B5563",
            othermonthwebackground=_BG, othermonthweforeground="#4B5563",
            showweeknumbers=False, font=("Segoe UI", 10),
            headerfont=("Segoe UI", 10, "bold"),
        )
        try:
            cal_kwargs["locale"] = "ru_RU"
            import babel  # noqa
        except Exception:
            cal_kwargs.pop("locale", None)

        cals_frame = ctk.CTkFrame(self, fg_color=_CARD, corner_radius=10,
                                  border_width=1, border_color=_BORDER)
        cals_frame.pack(fill="x", padx=20, pady=(16, 8))

        for side, label_text, attr in [("left", "Начало периода", "_cal_start"),
                                        ("left", "Конец периода",  "_cal_end")]:
            fr = ctk.CTkFrame(cals_frame, fg_color="transparent")
            fr.pack(side=side, padx=(16 if attr == "_cal_start" else 8,
                                     8  if attr == "_cal_start" else 16), pady=12)
            ctk.CTkLabel(fr, text=label_text,
                         font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
                         text_color=_SUBLABEL).pack(pady=(0, 6))
            cal = Calendar(fr, **cal_kwargs)
            cal.pack()
            setattr(self, attr, cal)

        ctk.CTkButton(self, text="Применить период", height=46,
                      font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
                      fg_color=_RED, hover_color=_RED_HOVER,
                      text_color=_WHITE, corner_radius=12,
                      command=self._apply).pack(fill="x", padx=20, pady=(0, 16))

    def _apply(self):
        try:
            s = self._cal_start.selection_get()
            e = self._cal_end.selection_get()
        except Exception:
            return
        period = f"{s.day:02d}.{s.month:02d}-{e.day:02d}.{e.month:02d}"
        self.destroy()
        self._on_confirm(period)

# ─── Главное окно приложения ─────────────────────────────────────────────────

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.title("Цифры")
        self.geometry("780x620")
        self.resizable(False, False)
        self.configure(fg_color=_BG)

        self._cfg          = _load_config()
        self._data_folder: str | None = None
        self._hist_path: Path | None = None
        self._last_summary = None
        self._pages: dict[str, ctk.CTkFrame] = {}
        self._nav_btns: dict[str, ctk.CTkButton] = {}

        self._build_header()
        self._build_nav()
        self._build_page_cifry()
        self._build_page_dinamika()
        self._build_page_top()
        self._show_page("cifry")

        saved = self._cfg.get("last_folder", "")
        if saved and Path(saved).exists():
            self._folder_var.set(saved)
            self._set_data_folder(saved)
        self._period_var.set(_auto_period())

    # ── Шапка ────────────────────────────────────────────────────────

    def _build_header(self):
        header = ctk.CTkFrame(self, fg_color=_DARK, corner_radius=0, height=64)
        header.pack(fill="x"); header.pack_propagate(False)

        logo_box = ctk.CTkFrame(header, fg_color=_RED, corner_radius=8, width=36, height=36)
        logo_box.pack(side="left", padx=(20, 12), pady=14)
        logo_box.pack_propagate(False)
        ctk.CTkLabel(logo_box, text="Я",
                     font=ctk.CTkFont(family="Segoe UI", size=17, weight="bold"),
                     text_color=_WHITE).place(relx=0.5, rely=0.5, anchor="center")

        col = ctk.CTkFrame(header, fg_color="transparent")
        col.pack(side="left")
        ctk.CTkLabel(col, text="Цифры",
                     font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
                     text_color=_WHITE, anchor="w").pack(anchor="w")
        ctk.CTkLabel(col, text="Автосборка еженедельного отчёта",
                     font=ctk.CTkFont(family="Segoe UI", size=10),
                     text_color="#6B7280", anchor="w").pack(anchor="w")

    # ── Навигация ─────────────────────────────────────────────────────

    def _build_nav(self):
        nav = ctk.CTkFrame(self, fg_color="#0E0E0E", corner_radius=0, height=46,
                           border_width=0)
        nav.pack(fill="x"); nav.pack_propagate(False)

        ctk.CTkFrame(nav, fg_color=_BORDER, height=1, corner_radius=0).pack(
            side="bottom", fill="x")

        for key, label in [("cifry", "Цифры"), ("dinamika", "Динамика"),
                            ("top", "Топ / Анти Топ")]:
            btn = ctk.CTkButton(
                nav, text=label, width=140, height=38,
                fg_color="transparent", hover_color="#1E1E1E",
                text_color="#555555", corner_radius=0,
                font=ctk.CTkFont(family="Segoe UI", size=13),
                command=lambda k=key: self._show_page(k),
            )
            btn.pack(side="left", padx=(4, 0))
            self._nav_btns[key] = btn

    def _show_page(self, key: str):
        for k, frame in self._pages.items():
            frame.pack_forget()
        self._pages[key].pack(fill="both", expand=True)
        for k, btn in self._nav_btns.items():
            if k == key:
                btn.configure(fg_color=_BG, text_color=_RED,
                               font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"))
            else:
                btn.configure(fg_color="transparent", text_color="#555555",
                               font=ctk.CTkFont(family="Segoe UI", size=13))
        if key == "dinamika":
            self._din_refresh_employees()
        if key == "top":
            self._top_refresh()

    # ── Вкладка: Цифры ────────────────────────────────────────────────

    def _build_page_cifry(self):
        page = ctk.CTkFrame(self, fg_color=_BG, corner_radius=0)
        self._pages["cifry"] = page

        card = ctk.CTkFrame(page, fg_color=_CARD, corner_radius=14,
                            border_width=1, border_color=_BORDER)
        card.pack(fill="both", expand=True, padx=24, pady=(16, 10))

        self._lbl(card, "ПАПКА С ДАННЫМИ")
        row1 = ctk.CTkFrame(card, fg_color="transparent")
        row1.pack(fill="x", padx=24, pady=(0, 14))
        self._folder_var = ctk.StringVar()
        self._folder_entry = ctk.CTkEntry(
            row1, textvariable=self._folder_var,
            placeholder_text="Выберите папку с выгрузками DataLens…",
            height=40, border_color=_BORDER, border_width=1,
            fg_color=_FIELD, text_color=_LABEL, corner_radius=10,
            font=ctk.CTkFont(family="Segoe UI", size=13),
        )
        self._folder_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkButton(row1, text="Выбрать", width=84, height=40,
                      fg_color=_GHOST, hover_color=_GHOST_HOV,
                      text_color=_LABEL, corner_radius=10,
                      border_width=1, border_color=_BORDER,
                      font=ctk.CTkFont(family="Segoe UI", size=13),
                      command=self._pick_folder).pack(side="left")

        ctk.CTkFrame(card, fg_color=_BORDER, height=1, corner_radius=0).pack(fill="x", padx=24)

        self._lbl(card, "ОТЧЁТНЫЙ ПЕРИОД")
        row2 = ctk.CTkFrame(card, fg_color="transparent")
        row2.pack(fill="x", padx=24, pady=(0, 20))
        self._period_var = ctk.StringVar()
        self._period_entry = ctk.CTkEntry(
            row2, textvariable=self._period_var,
            placeholder_text="например: 19.05–25.05",
            height=40, border_color=_BORDER, border_width=1,
            fg_color=_FIELD, text_color=_LABEL, corner_radius=10,
            font=ctk.CTkFont(family="Segoe UI", size=13),
        )
        self._period_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self._period_entry.bind("<Return>", lambda e: self._run())
        ctk.CTkButton(row2, text="📅", width=40, height=40,
                      fg_color=_GHOST, hover_color=_GHOST_HOV,
                      text_color=_LABEL, corner_radius=10,
                      border_width=1, border_color=_BORDER,
                      font=ctk.CTkFont(family="Segoe UI", size=15),
                      command=self._open_date_picker).pack(side="left")

        self._run_btn = ctk.CTkButton(
            card, text="Сформировать отчёт", height=50,
            font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
            fg_color=_RED, hover_color=_RED_HOVER,
            text_color=_WHITE, corner_radius=12,
            command=self._run,
        )
        self._run_btn.pack(fill="x", padx=24, pady=(0, 18))

        self._progress = ctk.CTkProgressBar(card, height=4, corner_radius=2,
                                            progress_color=_RED, fg_color="#272727")
        self._progress.set(0)
        self._progress.pack(fill="x", padx=24, pady=(0, 5))
        self._status_lbl = ctk.CTkLabel(card, text="", height=16,
                                        font=ctk.CTkFont(family="Segoe UI", size=11),
                                        text_color=_SUBLABEL, anchor="w")
        self._status_lbl.pack(fill="x", padx=24, pady=(0, 12))

        self._result_frame = ctk.CTkFrame(card, fg_color="transparent")
        self._result_frame.pack(fill="x", padx=24, pady=(0, 16))
        self._result_lbl = ctk.CTkLabel(self._result_frame, text="",
                                        font=ctk.CTkFont(family="Segoe UI", size=12),
                                        text_color=_LABEL, anchor="w")
        self._result_lbl.pack(side="left", fill="x", expand=True)
        self._hist_btn = ctk.CTkButton(
            self._result_frame, text="История", width=82, height=32,
            fg_color=_GHOST, hover_color=_GHOST_HOV, text_color=_LABEL,
            corner_radius=8, border_width=1, border_color=_BORDER,
            font=ctk.CTkFont(family="Segoe UI", size=12),
            command=self._open_history)

        ctk.CTkLabel(page, text="Цифры by Тимур",
                     font=ctk.CTkFont(family="Segoe UI", size=10),
                     text_color="#D1D5DB").pack(pady=(0, 8))

    # ── Вкладка: Динамика ─────────────────────────────────────────────

    def _build_page_dinamika(self):
        page = ctk.CTkFrame(self, fg_color=_BG, corner_radius=0)
        self._pages["dinamika"] = page

        ctrl_card = ctk.CTkFrame(page, fg_color=_CARD, corner_radius=14,
                                 border_width=1, border_color=_BORDER)
        ctrl_card.pack(fill="x", padx=24, pady=(16, 8))

        row = ctk.CTkFrame(ctrl_card, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=14)

        ctk.CTkLabel(row, text="Сотрудник",
                     font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
                     text_color="#9CA3AF").pack(side="left", padx=(0, 6))
        self._din_emp_var = ctk.StringVar(value="—")
        self._din_emp_menu = ctk.CTkOptionMenu(
            row, variable=self._din_emp_var,
            values=["—"], width=180, height=36,
            fg_color=_FIELD, button_color=_RED, button_hover_color=_RED_HOVER,
            text_color=_LABEL, font=ctk.CTkFont(family="Segoe UI", size=12),
            command=lambda _: self._draw_chart(),
        )
        self._din_emp_menu.pack(side="left", padx=(0, 18))

        ctk.CTkLabel(row, text="Показатель",
                     font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
                     text_color="#9CA3AF").pack(side="left", padx=(0, 6))
        metric_names = [cfg["name"] for cfg in METRICS]
        self._din_metric_var = ctk.StringVar(value=metric_names[0])
        ctk.CTkOptionMenu(
            row, variable=self._din_metric_var,
            values=metric_names, width=220, height=36,
            fg_color=_FIELD, button_color=_RED, button_hover_color=_RED_HOVER,
            text_color=_LABEL, font=ctk.CTkFont(family="Segoe UI", size=12),
            command=lambda _: self._draw_chart(),
        ).pack(side="left")

        # Matplotlib canvas
        chart_card = ctk.CTkFrame(page, fg_color=_CARD, corner_radius=14,
                                  border_width=1, border_color=_BORDER)
        chart_card.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        self._din_placeholder = ctk.CTkLabel(
            chart_card,
            text="Выбери папку с данными во вкладке «Цифры»\nчтобы загрузить историю",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=_SUBLABEL,
        )

        self._fig = Figure(figsize=(7, 4), dpi=96, facecolor="#1A1A1A")
        self._canvas = FigureCanvasTkAgg(self._fig, master=chart_card)
        self._canvas_widget = self._canvas.get_tk_widget()

    def _din_refresh_employees(self):
        if self._hist_path is None or not self._hist_path.exists():
            self._canvas_widget.pack_forget()
            self._din_placeholder.place(relx=0.5, rely=0.5, anchor="center")
            return
        self._din_placeholder.place_forget()
        self._canvas_widget.pack(fill="both", expand=True, padx=8, pady=8)

        periods = get_history_periods(self._hist_path)
        if not periods:
            return

        wb = openpyxl.load_workbook(self._hist_path, read_only=True, data_only=True)
        ws = wb.active
        metric_names = {cfg["name"] for cfg in METRICS}
        seen = set()
        employees = []
        for r in range(3, (ws.max_row or 0) + 1):
            v = ws.cell(row=r, column=1).value
            if v:
                s = str(v).strip()
                if s and s not in metric_names and s not in seen:
                    seen.add(s)
                    employees.append(s)
        wb.close()

        if employees:
            self._din_emp_menu.configure(values=employees)
            if self._din_emp_var.get() not in employees:
                self._din_emp_var.set(employees[0])
        self._draw_chart()

    def _draw_chart(self):
        if self._hist_path is None or not self._hist_path.exists():
            return
        emp    = self._din_emp_var.get()
        metric = self._din_metric_var.get()
        series = load_history_for_chart(self._hist_path, emp, metric)

        self._fig.clear()
        ax = self._fig.add_subplot(111)
        ax.set_facecolor("#1A1A1A")
        for spine in ax.spines.values():
            spine.set_edgecolor("#272727")
        ax.tick_params(colors="#666666", labelsize=9)
        ax.yaxis.set_major_formatter(ticker.FormatStrFormatter("%.2f"))

        if series.empty:
            ax.text(0.5, 0.5, "Нет данных для отображения",
                    ha="center", va="center", fontsize=12, color="#555555",
                    transform=ax.transAxes)
        else:
            periods = list(series.index)
            values  = list(series.values)
            x = range(len(periods))
            ax.plot(x, values, color=_RED, linewidth=2.5,
                    marker="o", markersize=7, markerfacecolor=_WHITE,
                    markeredgecolor=_RED, markeredgewidth=2)
            ax.fill_between(x, values, alpha=0.08, color=_RED)
            ax.set_xticks(list(x))
            ax.set_xticklabels(periods, rotation=25, ha="right", fontsize=9)
            ax.set_title(f"{emp}  ·  {metric}", fontsize=11,
                         color="#F0F0F0", pad=10, fontweight="bold")
            ax.grid(axis="y", color="#272727", linewidth=1)
            ax.set_axisbelow(True)
            # Подписи значений
            for xi, yi in zip(x, values):
                ax.annotate(f"{yi:.2f}", xy=(xi, yi),
                            xytext=(0, 10), textcoords="offset points",
                            ha="center", fontsize=8, color="#888888")

        self._fig.tight_layout()
        self._canvas.draw()

    # ── Вкладка: Топ / Анти Топ ───────────────────────────────────────

    def _build_page_top(self):
        page = ctk.CTkFrame(self, fg_color=_BG, corner_radius=0)
        self._pages["top"] = page

        hdr_row = ctk.CTkFrame(page, fg_color="transparent")
        hdr_row.pack(fill="x", padx=24, pady=(16, 8))
        ctk.CTkLabel(hdr_row, text="Топ и Анти Топ",
                     font=ctk.CTkFont(family="Segoe UI", size=15, weight="bold"),
                     text_color=_LABEL).pack(side="left")
        self._top_period_lbl = ctk.CTkLabel(hdr_row, text="На основе последнего периода в Истории",
                     font=ctk.CTkFont(family="Segoe UI", size=11),
                     text_color=_SUBLABEL)
        self._top_period_lbl.pack(side="left", padx=(10, 0))
        ctk.CTkButton(hdr_row, text="Обновить", width=100, height=34,
                      fg_color=_GHOST, hover_color=_GHOST_HOV,
                      text_color=_LABEL, corner_radius=8,
                      border_width=1, border_color=_BORDER,
                      font=ctk.CTkFont(family="Segoe UI", size=12),
                      command=self._top_refresh).pack(side="right")

        self._top_placeholder = ctk.CTkLabel(
            page,
            text="История пока пуста.\nСформируй отчёт во вкладке «Цифры».",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=_SUBLABEL,
        )

        self._top_scroll = ctk.CTkScrollableFrame(
            page, fg_color=_BG, corner_radius=0,
        )

    def _top_refresh(self):
        if self._hist_path is None or not self._hist_path.exists():
            self._top_scroll.pack_forget()
            self._top_placeholder.place(relx=0.5, rely=0.55, anchor="center")
            return

        summary = load_last_period_summary(self._hist_path)
        if summary is None:
            self._top_scroll.pack_forget()
            self._top_placeholder.place(relx=0.5, rely=0.55, anchor="center")
            return

        periods = get_history_periods(self._hist_path)
        last_period = periods[-1] if periods else "—"
        self._top_period_lbl.configure(text=f"Период: {last_period}")

        self._top_placeholder.place_forget()
        self._top_scroll.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        for w in self._top_scroll.winfo_children():
            w.destroy()

        GREEN_HDR = "#14532D"
        RED_HDR   = "#7F1D1D"
        GREEN_ROW = "#0D2018"
        RED_ROW   = "#200D0D"

        for metric in _RANKING_METRICS:
            if metric not in summary.columns:
                continue

            col_data  = summary[metric].dropna()
            top_list  = sorted(
                [(n, v) for n, v in col_data.items() if _TOP_CRITERIA[metric](v)],
                key=lambda x: x[1], reverse=_TOP_SORT_REVERSE[metric],
            )
            anti_list = sorted(
                [(n, v) for n, v in col_data.items() if _ANTI_TOP_CRITERIA[metric](v)],
                key=lambda x: x[1], reverse=_ANTI_TOP_SORT_REVERSE[metric],
            )

            # Блок метрики
            block = ctk.CTkFrame(self._top_scroll, fg_color=_CARD,
                                 corner_radius=12, border_width=1, border_color=_BORDER)
            block.pack(fill="x", pady=(0, 12))

            # Заголовок метрики
            mhdr = ctk.CTkFrame(block, fg_color=_RED, corner_radius=0,
                                height=36)
            mhdr.pack(fill="x"); mhdr.pack_propagate(False)
            ctk.CTkLabel(mhdr, text=metric,
                         font=ctk.CTkFont(family="Segoe UI", size=13, weight="bold"),
                         text_color=_WHITE).pack(side="left", padx=14)

            cols_frame = ctk.CTkFrame(block, fg_color="transparent")
            cols_frame.pack(fill="x", padx=12, pady=10)
            cols_frame.columnconfigure(0, weight=1)
            cols_frame.columnconfigure(1, weight=0)
            cols_frame.columnconfigure(2, weight=1)

            # Колонка ТОП
            top_col = ctk.CTkFrame(cols_frame, fg_color="transparent")
            top_col.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
            top_hdr = ctk.CTkFrame(top_col, fg_color=GREEN_HDR, corner_radius=8, height=30)
            top_hdr.pack(fill="x"); top_hdr.pack_propagate(False)
            ctk.CTkLabel(top_hdr,
                         text=f"ТОП — {len(top_list)} чел.   (критерий: {_TOP_CRITERIA_TEXT[metric]})",
                         font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
                         text_color=_WHITE).pack(side="left", padx=10)
            for i, (name, val) in enumerate(top_list):
                disp = int(val) if isinstance(val, float) and val == int(val) else round(val, 2)
                row_f = ctk.CTkFrame(top_col, fg_color=GREEN_ROW if i%2==0 else _CARD,
                                     corner_radius=0, height=28)
                row_f.pack(fill="x"); row_f.pack_propagate(False)
                ctk.CTkLabel(row_f, text=name,
                             font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
                             text_color="#6EE7A0").pack(side="left", padx=8)
                ctk.CTkLabel(row_f, text=str(disp),
                             font=ctk.CTkFont(family="Segoe UI", size=11),
                             text_color="#6EE7A0").pack(side="right", padx=8)

            # Разделитель
            ctk.CTkFrame(cols_frame, fg_color=_BORDER, width=1,
                         corner_radius=0).grid(row=0, column=1, sticky="ns", padx=4)

            # Колонка Анти ТОП
            anti_col = ctk.CTkFrame(cols_frame, fg_color="transparent")
            anti_col.grid(row=0, column=2, sticky="nsew", padx=(6, 0))
            anti_hdr = ctk.CTkFrame(anti_col, fg_color=RED_HDR, corner_radius=8, height=30)
            anti_hdr.pack(fill="x"); anti_hdr.pack_propagate(False)
            ctk.CTkLabel(anti_hdr,
                         text=f"Анти ТОП — {len(anti_list)} чел.   (критерий: {_ANTI_TOP_CRITERIA_TEXT[metric]})",
                         font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
                         text_color=_WHITE).pack(side="left", padx=10)
            for i, (name, val) in enumerate(anti_list):
                disp = int(val) if isinstance(val, float) and val == int(val) else round(val, 2)
                row_f = ctk.CTkFrame(anti_col, fg_color=RED_ROW if i%2==0 else _CARD,
                                     corner_radius=0, height=28)
                row_f.pack(fill="x"); row_f.pack_propagate(False)
                ctk.CTkLabel(row_f, text=name,
                             font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
                             text_color="#FCA5A5").pack(side="left", padx=8)
                ctk.CTkLabel(row_f, text=str(disp),
                             font=ctk.CTkFont(family="Segoe UI", size=11),
                             text_color="#FCA5A5").pack(side="right", padx=8)

    # ── Общие методы ──────────────────────────────────────────────────

    @staticmethod
    def _lbl(parent, text: str):
        ctk.CTkLabel(parent, text=text,
                     font=ctk.CTkFont(family="Segoe UI", size=10, weight="bold"),
                     text_color="#555555", anchor="w").pack(fill="x", padx=24, pady=(16, 5))

    def _set_data_folder(self, folder: str):
        self._data_folder = folder
        hist = Path(folder).parent / HISTORY_FILE_NAME
        self._hist_path = hist if hist.exists() else None

    def _pick_folder(self):
        from tkinter import filedialog
        folder = filedialog.askdirectory(title="Выберите папку с выгрузками")
        if folder:
            self._folder_var.set(folder)
            self._folder_entry.configure(border_color=_BORDER)
            self._set_data_folder(folder)
            self._cfg["last_folder"] = folder
            _save_config(self._cfg)

    def _open_date_picker(self):
        DatePickerWindow(self, on_confirm=lambda p: self._period_var.set(p))

    def _flash_error(self, entry, msg: str):
        entry.configure(border_color="#DC2626")
        self._show_warn(msg)
        self.after(2500, lambda: entry.configure(border_color=_BORDER))

    def _run(self):
        folder = self._folder_var.get().strip()
        period = self._period_var.get().strip()
        if not folder:
            self._flash_error(self._folder_entry, "Укажи папку с данными")
            return
        if not Path(folder).exists():
            self._flash_error(self._folder_entry, "Папка не найдена — проверь путь")
            return
        if not period:
            self._flash_error(self._period_entry, "Введи период или выбери через календарь")
            return

        self._run_btn.configure(state="disabled", text="Загружаю список…")
        self._result_lbl.configure(text="")
        self._hist_btn.pack_forget()

        def load_and_show():
            try:
                result = scan_folder(folder)
                self.after(0, self._show_filter, folder, period, result)
            except Exception as exc:
                self.after(0, self._on_error, str(exc))

        threading.Thread(target=load_and_show, daemon=True).start()

    def _show_filter(self, folder, period, scan_result):
        self._run_btn.configure(state="normal", text="Сформировать отчёт")
        excluded = set(self._cfg.get("excluded_employees", []))
        FilterWindow(
            self, employees=scan_result["employees"],
            scan_result=scan_result, excluded=excluded,
            on_confirm=lambda sel, fired: self._start(folder, period, sel, fired),
        )

    def _start(self, folder, period, employee_filter, fired):
        if not employee_filter:
            self._show_warn("Нет выбранных сотрудников")
            return
        self._run_btn.configure(state="disabled", text="Обрабатываю…")
        self._progress.set(0)
        self._status_lbl.configure(text="Запуск…")
        self._out_path = None
        threading.Thread(
            target=self._worker, args=(folder, period, employee_filter, fired), daemon=True
        ).start()

    def _worker(self, folder, period, employee_filter, fired):
        try:
            out = build_report(
                folder, period, employee_filter, fired,
                progress_cb=lambda v: self.after(0, self._progress.set, v),
                status_cb=lambda s: self.after(0, self._status_lbl.configure, {"text": s}),
            )
            self._out_path = out
            self.after(0, self._on_success, out)
        except Exception as exc:
            self.after(0, self._on_error, str(exc))

    def _on_success(self, hist_path: str):
        self._run_btn.configure(state="normal", text="Сформировать отчёт")
        self._progress.set(1)
        self._status_lbl.configure(text="")
        hist = Path(hist_path)
        if hist.exists():
            self._hist_path = hist
            self._result_lbl.configure(text="✓  История обновлена", text_color="#059669")
            self._hist_btn.pack(side="right")
            _open_path(hist)
        else:
            self._result_lbl.configure(text="✓  Отчёт сформирован", text_color="#059669")
        self._din_refresh_employees()
        _notify("Цифры — отчёт готов", "История обновлена")

    def _on_error(self, msg: str):
        self._run_btn.configure(state="normal", text="Сформировать отчёт")
        self._progress.set(0)
        self._status_lbl.configure(text="")
        self._result_lbl.configure(text=f"✗  {msg}", text_color="#DC2626")

    def _show_warn(self, msg: str):
        self._result_lbl.configure(text=f"  {msg}", text_color="#D97706")
        self._hist_btn.pack_forget()

    def _open_history(self):
        if self._hist_path and self._hist_path.exists():
            HistoryManagerWindow(self, self._hist_path)


# ─── Точка входа ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
